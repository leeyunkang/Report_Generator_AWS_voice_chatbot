import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta
import pandas as pd
from numpy import nan
from io import BytesIO
import pytz
import subprocess
import os


def query_table(db_client, TABLE_NAME_1, start_timestamp, end_timestamp):
  # QUERY ANSWERED CALLS
  answered_response_items = []
  answered_response = db_client.query(
      TableName=TABLE_NAME_1,
      IndexName="Call_Answered-index",
      ExpressionAttributeValues={
          ":ca": {"S": "True"},
          ":start_ts": {"S": start_timestamp},
          ":end_ts": {"S": end_timestamp}
      },
      KeyConditionExpression="Call_Answered = :ca",
      FilterExpression="Trigger_Timestamp BETWEEN :start_ts AND :end_ts",
  )
  answered_response_items.extend(answered_response.get('Items', []))

  # Continue querying while paginated results exist
  while 'LastEvaluatedKey' in answered_response:
      answered_response = db_client.query(
          TableName=TABLE_NAME_1,
          IndexName="Call_Answered-index",
          ExpressionAttributeValues={
              ":ca": {"S": "True"},
              ":start_ts": {"S": start_timestamp},
              ":end_ts": {"S": end_timestamp}
          },
          KeyConditionExpression="Call_Answered = :ca",
          FilterExpression="Trigger_Timestamp BETWEEN :start_ts AND :end_ts",
          ExclusiveStartKey=answered_response['LastEvaluatedKey']
      )
      answered_response_items.extend(answered_response.get('Items', []))

  # QUERY UNANSWERED CALLS
  unanswered_response_items = []
  unanswered_response = db_client.query(
      TableName=TABLE_NAME_1,
      IndexName="Call_Answered-index",
      ExpressionAttributeValues={
          ":ca": {"S": "False"},
          ":start_ts": {"S": start_timestamp},
          ":end_ts": {"S": end_timestamp}
      },
      KeyConditionExpression="Call_Answered = :ca",
      FilterExpression="Trigger_Timestamp BETWEEN :start_ts AND :end_ts",
  )
  unanswered_response_items.extend(unanswered_response.get('Items', []))

  # Continue querying while paginated results exist
  while 'LastEvaluatedKey' in unanswered_response:
      unanswered_response = db_client.query(
          TableName=TABLE_NAME_1,
          IndexName="Call_Answered-index",
          ExpressionAttributeValues={
              ":ca": {"S": "False"},
              ":start_ts": {"S": start_timestamp},
              ":end_ts": {"S": end_timestamp}
          },
          KeyConditionExpression="Call_Answered = :ca",
          FilterExpression="Trigger_Timestamp BETWEEN :start_ts AND :end_ts",
          ExclusiveStartKey=unanswered_response['LastEvaluatedKey']
      )
      unanswered_response_items.extend(unanswered_response.get('Items', []))

  return answered_response_items, unanswered_response_items


def answered_calls(response_items):
  answered_response_df = []

  for item in response_items:
    policy_number = item.get('Policy_Number', {}).get('S', None)
    if policy_number.__contains__("HANATEST"):
      continue # Skip rows with "HANATEST"
    last_stage = next(iter(item.get('Bot_Comprehensibility', {})["L"][-1]["M"]))
    entity = classify_entity(item.get('Policy_Number', {}).get('S', ''))
    verification_status = classify_verification_status(item.get('Verification', []))
    policy_received = classify_policy_received(item.get('Policy_Received'))
    survey_rating = classify_survey_rating(item.get('Survey_Rating'))
    hana_call_time = item.get('Trigger_Timestamp', {}).get('S', None)
    phone_number = item.get('Policyholder_Phone_Number', {}).get('S', None)

    answered_response_df.append({
        "Policy_Number": policy_number,
        "Entity": entity,
        "Phone_Number": phone_number,
        "HANA Call Time": hana_call_time,
        "Verification": verification_status,
        "Policy Received": policy_received,
        "Survey Rating": survey_rating,
        "Last Stage": last_stage,
    })

  # Create DataFrame from the list of items
  answered_df1 = pd.DataFrame(answered_response_df)

  return answered_df1


def unanswered_calls(response_items):
  response_df = []
  for item in response_items:
    unanswered_policy_number = item.get('Policy_Number', {}).get('S', None)
    if unanswered_policy_number.__contains__("HANATEST"):
      continue # Skip rows with "HANATEST"

    unanswered_phone_number = item.get('Policyholder_Phone_Number', {}).get('S', None)

    entity = classify_entity(unanswered_policy_number)

    response_df.append({
          "Policy_Number": unanswered_policy_number,
          "Entity": entity,
          "Phone_Number": unanswered_phone_number,
      })
  df2 = pd.DataFrame(response_df)
  df2 = df2.drop_duplicates("Policy_Number", keep='first')

  return df2

def classify_entity(policy_number):
    if policy_number.startswith("TR") or policy_number.startswith("LR"):
        return "FTA"
    else:
        return "LIA"


def classify_verification_status(verification):
    if not verification:
        return "NA"
    last_key = next(iter(verification["L"][-1]["M"]))
    return "PASSED" if verification["L"][-1]["M"][last_key]["S"] == "True" else "FAILED"


def classify_policy_received(policy_received):
    if not policy_received:
        return "Did not reach the stage"
    return "YES" if policy_received["S"] == "True" else "NO"


def classify_survey_rating(survey_rating):
    if not survey_rating:
        return "Not applicable"
    return [{key: value['S']} for ratings in survey_rating["L"] for key, value in ratings['M'].items()]

def clean_data(db_resource, TABLE_NAME_2, BUCKET_NAME, s3_client, df1, df2,date_1,date_2,date_3,times):
  stages = ["NA",nan, 'T.1', 'F.1', '1.1', '1.2', '1.3', '2.1', '2.2', '2.3', '3.1', '3.2', '4.1', '4.2', '4.3', '5.1', '5.2', '5.3', '5.4', '5.5', '6.1']
  ranking_dict = {i: stages.index(i) for i in stages}
  df1.loc[:, "Stages_Reached"] = df1["Last Stage"].apply(lambda x: ranking_dict[x])
  df1 = df1.sort_values(by=["HANA Call Time", "Stages_Reached"],ascending=False)
  df1 = df1.drop_duplicates(subset=["Policy_Number"],keep="first")

  filtered_df1 = df1[df1['Last Stage'] == 'T.1']
  filtered_df1 = filtered_df1[['Policy_Number', 'Entity', 'Phone_Number']]
  df1.drop(filtered_df1.index, inplace=True)

  df1['HANA Call Time'] = pd.to_datetime(df1['HANA Call Time'])
  df1['HANA Call Time'] = df1['HANA Call Time'].dt.strftime('%H:%M')

  df1 = df1.sort_values(by=["HANA Call Time"],ascending=True)
  df1 = df1.drop(columns=["Stages_Reached"],axis=1)

  concatenated_unanswered_df = pd.concat([df2, filtered_df1], axis=0)
  filtered_unanswered = concatenated_unanswered_df[~concatenated_unanswered_df["Policy_Number"].isin(df1["Policy_Number"])]
  clean_unanswered = filtered_unanswered.drop_duplicates('Policy_Number')

  date1 = date_1.split("T")[0]
  date2 =date_2.split("T")[0]
  date3 = date_3.split("T")[0]


  if times == 0:
     current_date_str = date1
     tmr_str = date2
     two_days_after_str = date3
  elif times == 2 :
     day_before_str = date1
     current_date_str = date2
     tmr_str = date3
  else :
     day_before_str = date2
     current_date_str = date3
     tmr_str = None

    # Specify the bucket name and the object (file) key
  bucket_name = BUCKET_NAME
  if times==0:
    object_key = 'Hana Call Summary template/Hana Call Summary Full Report-Template.xlsx'
  elif times==2:
    object_key = f'Hana_Call_Summary_Report/Hana Call Summary Full Report {day_before_str}.xlsx'
  else :
    object_key = f'Hana_Call_Summary_Report/Hana Call Summary Full Report {day_before_str}.xlsx'

  # Download the Excel file into memory as bytes
  response = s3_client.get_object(Bucket=bucket_name, Key=object_key)

  excel_bytes = response['Body'].read()
  workbook = openpyxl.load_workbook(BytesIO(excel_bytes))
  #delete the last sheet
  sheet_names = workbook.sheetnames
  if sheet_names:
      # Delete the last sheet
      last_sheet_name = sheet_names[-1]
      workbook.remove(workbook[last_sheet_name])

  # Define the table for resource
  table = db_resource.Table(TABLE_NAME_2)

  if tmr_str is not None:
    try:
        tmr_str_reschedule = datetime.strptime(tmr_str, '%d-%m-%Y')
        formatted_date = tmr_str_reschedule.strftime('%Y-%m-%d')

        # Get the current date and time in the 'Asia/Singapore' timezone
        tz = pytz.timezone('Asia/Singapore')
        current_datetime = datetime.now(tz)

        # Calculate tomorrow's date
        tomorrow = current_datetime + timedelta(days=1)

        # Scan the table  items with matching Schedule_ID and today's date
        response = table.scan(
          FilterExpression="Schedule_ID = :schedule_id and begins_with(Schedule_Call_Timestamp, :date) and (begins_with(Policy_Number, :policy_cr) or begins_with(Policy_Number, :policy_ir))",
          ExpressionAttributeValues={":schedule_id": "Rescheduled", ":date": formatted_date, ":policy_cr": "CR", ":policy_ir": "IR"}
      )

        count_rereschedule_call_lia = response['Count']

        # Scan the table items with matching Schedule_ID and today's date
        response = table.scan(
          FilterExpression="Schedule_ID = :schedule_id and begins_with(Schedule_Call_Timestamp, :date) and (begins_with(Policy_Number, :policy_lr) or begins_with(Policy_Number, :policy_tr))",
          ExpressionAttributeValues={":schedule_id": "Rescheduled", ":date": formatted_date, ":policy_lr": "LR", ":policy_tr": "TR"}
      )


        count_rereschedule_call_fta = response['Count']

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        count_rereschedule_call_fta = 0
        count_rereschedule_call_lia = 0

  count_ofta_answered = df1["Entity"].value_counts().get("FTA", 0)
  count_olia_answered = df1["Entity"].value_counts().get("LIA", 0)

  count_ofta_unanswered =clean_unanswered["Entity"].value_counts().get("FTA", 0)
  count_olia_unanswered =clean_unanswered["Entity"].value_counts().get("LIA", 0)

  total_call_ofta = count_ofta_answered + count_ofta_unanswered
  total_call_olia = count_olia_answered + count_olia_unanswered



  # Specify the sheet want to work with
  sheet_name_first = 'Key Highlights'
  sheet_first = workbook[sheet_name_first]

  if times == 0 :


    cell = sheet_first['B6']
    cell.value = current_date_str


    cell = sheet_first['E6']
    cell.value = total_call_ofta

    cell = sheet_first['E7']
    cell.value = '0'
    cell = sheet_first['E23']
    cell.value = count_rereschedule_call_fta

    cell = sheet_first['E8']
    cell.value = count_ofta_answered
    cell = sheet_first['E9']
    cell.value = count_ofta_unanswered

    cell = sheet_first['E10']
    cell.value = total_call_olia

    cell = sheet_first['E11']
    cell.value = '0'
    cell = sheet_first['E27']
    cell.value = count_rereschedule_call_lia

    cell = sheet_first['E12']
    cell.value = count_olia_answered
    cell = sheet_first['E13']
    cell.value = count_olia_unanswered

    cell = sheet_first['E20']
    cell.value = count_ofta_unanswered

    cell = sheet_first['E26']
    cell.value = count_olia_unanswered

    #table 2
    cell = sheet_first['B20']
    cell.value = tmr_str


    cell = sheet_first['E24']
    cell.value = "Will be updated "+tmr_str
    cell = sheet_first['E25']
    cell.value = "Will be updated "+tmr_str

    cell = sheet_first['E28']
    cell.value = "Will be updated "+tmr_str
    cell = sheet_first['E29']
    cell.value = "Will be updated "+tmr_str
    #table 3
    cell = sheet_first['B36']
    cell.value = two_days_after_str

    cell = sheet_first['E36']
    cell.value = "Will be updated "+tmr_str
    cell = sheet_first['E40']
    cell.value =  "Will be updated "+tmr_str
    cell = sheet_first['E41']
    cell.value =  "Will be updated "+two_days_after_str
    cell = sheet_first['E42']
    cell.value =  "Will be updated "+two_days_after_str


    cell = sheet_first['E43']
    cell.value =  "Will be updated "+tmr_str
    cell = sheet_first['E44']
    cell.value =  "Will be updated "+tmr_str
    cell = sheet_first['E45']
    cell.value =  "Will be updated "+two_days_after_str
    cell = sheet_first['E46']
    cell.value =  "Will be updated "+two_days_after_str

  elif times == 2 :

    cell = sheet_first['E40']
    cell.value = count_rereschedule_call_fta

    cell = sheet_first['E24']
    cell.value = count_ofta_answered
    cell = sheet_first['E25']
    cell.value = count_ofta_unanswered

    cell = sheet_first['E44']
    cell.value = count_rereschedule_call_lia

    cell = sheet_first['E28']
    cell.value = count_olia_answered
    cell = sheet_first['E29']
    cell.value = count_olia_unanswered

    cell = sheet_first['E36']
    cell.value = count_ofta_unanswered
    cell = sheet_first['E43']
    cell.value = count_olia_unanswered

    #table 3


    cell = sheet_first['E41']
    cell.value =  "Will be updated "+tmr_str
    cell = sheet_first['E42']
    cell.value =  "Will be updated "+tmr_str



    cell = sheet_first['E45']
    cell.value =  "Will be updated "+tmr_str
    cell = sheet_first['E46']
    cell.value =  "Will be updated "+tmr_str

  else:


    cell = sheet_first['E41']
    cell.value = count_ofta_answered
    cell = sheet_first['E42']
    cell.value = count_ofta_unanswered


    cell = sheet_first['E45']
    cell.value = count_olia_answered
    cell = sheet_first['E46']
    cell.value = count_olia_unanswered


  # Create two new sheets
  new_sheet1 = workbook.create_sheet(title="Call Summary Run "+current_date_str)
  new_sheet2 = workbook.create_sheet(title="Unanswered "+current_date_str)

  #modify the call summary run
  sheet = workbook["Call Summary Run "+current_date_str]

  # Convert the date string to a datetime object
  date_object = datetime.strptime(current_date_str, "%d-%m-%Y")

  # Get the day of the week as a string
  day_of_week = date_object.strftime("%A")

  # Merge cells
  start_cell_3 = "B3"
  end_cell_3 = "J5"
  sheet.merge_cells(start_cell_3 + ":" + end_cell_3)

  # Set content in the merged cell
  sheet[start_cell_3] = "HANA Outbound Call Answered Summary - "+day_of_week+" , "+current_date_str

 # Apply formatting to the merged cell
  fill = PatternFill(start_color="ffdc64", end_color="ffdc64", fill_type="solid")
  font = Font(bold=True,underline="single")  # Bold and gold color
  alignment = Alignment(horizontal="center", vertical="center")


  # Apply the formatting to the merged cell
  for row in sheet.iter_rows(min_row=3, max_row=5, min_col=2, max_col=10):
      for cell in row:
          cell.fill = fill
          cell.font = font
          cell.alignment = alignment


  # Set the column widths
  for col_letter in ['A','B', 'C', 'D', 'E','F','G','H','I','J']:
      sheet.column_dimensions[col_letter].width = {
          'A':22.91,
          'B': 10.36,
          'C': 27.91,
          'D': 13.82,
          'E': 25.18,
          'F': 23.18,
          'G': 57.55,
          'H': 32.18,
          'I': 32.91,
          'J': 27.55
      }[col_letter]



  # Add a new line
  sheet.cell(row=6, column=2).value = ""  # Empty cell for a new line


  sheet.row_dimensions[7].height = 43.50

  # Add the header of table
  data_4 = [
      ["SI.No", "Policy_Number", "Entity", "Phone Number","HANA Call Time","NRIC                                                                                                               Verificatio Stage                                                                                                                       ( Did not reach the Stage / Passed / Fail)","Policy Document Received (Yes/No)","Survey Ratings","Dropped off Stage "]

  ]



  # Set formatting for the header of table
  font = Font(bold=True, color="000000")  # Black text for non-bold
  fill = PatternFill(start_color="fffc04", end_color="fffc04", fill_type="solid")
  border = Border(
      left=Side(style="thin"),
      right=Side(style="thin"),
      top=Side(style="thin"),
      bottom=Side(style="thin")
  )


  # Fill the header of table
  for row_index, row_data in enumerate(data_4, start=7):
      for col_index, value in enumerate(row_data, start=2):
          cell = sheet.cell(row=row_index, column=col_index)
          cell.value = value
          cell.font = font
          cell.border = border
          cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)  # Enable text wrapping
          if row_index == 7:
              cell.fill = fill


  ##sheet 3 unanswered call
  sheet = workbook["Unanswered "+current_date_str]

  # Merge cells
  start_cell_4 = "B2"
  end_cell_4 = "D2"
  sheet.merge_cells(start_cell_4 + ":" + end_cell_4)

  # Set content in the merged cell
  sheet[start_cell_4] = "Unanswered Calls"

  # Apply formatting to the merged cell
  fill = PatternFill(start_color="f07c34", end_color="f07c34", fill_type="solid")
  font = Font(bold=True)  # Bold
  alignment = Alignment(horizontal="center", vertical="center")


  # Apply the formatting to the merged cell
  for row in sheet.iter_rows(min_row=2, max_row=2, min_col=2, max_col=4):
      for cell in row:
          cell.fill = fill
          cell.font = font
          cell.alignment = alignment


  # Set the column widths
  for col_letter in ['A','B', 'C', 'D']:
      sheet.column_dimensions[col_letter].width = {
          'A':8.09,
          'B': 17.91,
          'C': 9.82,
          'D': 18.18

      }[col_letter]

  # Merge cells
  start_cell_5 = "B4"
  end_cell_5 = "D4"
  sheet.merge_cells(start_cell_5 + ":" + end_cell_5)

  # Set content in the merged cell
  sheet[start_cell_5] = current_date_str

  # Apply formatting to the merged cell
  fill = PatternFill(start_color="ffffff",end_color="ffffff",fill_type="solid")
  font = Font(bold=True)  # Bold
  alignment = Alignment(horizontal="center", vertical="center")
  border = Border(
      left=Side(style="thin"),
      right=Side(style="thin"),
      top=Side(style="thin"),
      bottom=Side(style="thin")
  )

  # Apply the formatting to the merged cell
  for row in sheet.iter_rows(min_row=4, max_row=4, min_col=2, max_col=4):
      for cell in row:
          cell.fill = fill
          cell.font = font
          cell.border = border
          cell.alignment = alignment

  data_4 = [
      ["Policy_Number", "Entity", "Phone_Number" ]

  ]


  # Set formatting for the header of table
  font = Font(bold=True, color="ffffff")  # white text for non-bold
  fill = PatternFill(start_color="8ea9db", end_color="8ea9db",fill_type="solid")
  border = Border(
      left=Side(style="thin"),
      right=Side(style="thin"),
      top=Side(style="thin"),
      bottom=Side(style="thin")
  )


  # Fill the second table data and apply formatting
  for row_index, row_data in enumerate(data_4, start=6):
      for col_index, value in enumerate(row_data, start=2):
          cell = sheet.cell(row=row_index, column=col_index)
          cell.value = value
          cell.font = font
          cell.border = border
          cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)  # Enable text wrapping
          if row_index == 6:
              cell.fill = fill


  # Create a new sheet or use an existing one to insert the DataFrame
  sheet_name = 'Call Summary Run '+current_date_str  # Specify the name of the sheet you want to insert the data into
  ws = workbook[sheet_name]

  alignment = Alignment(horizontal="center", vertical="center")
  border = Border(
      left=Side(style="thin"),
      right=Side(style="thin"),
      top=Side(style="thin"),
      bottom=Side(style="thin")
  )

  # create a row for the index number
  df1.insert(0, 'New_Column', range(1, len(df1) + 1))
  df1 = df1.astype(str)

  # Convert the DataFrame to a list of lists for inserting into the Excel sheet
  data_to_insert = [df1.columns.tolist()] + df1.values.tolist()

  # Determine the starting cell where you want to insert the data
  start_row = 8  # Specify the row where you want to start (1-based index)
  start_col = 2  # Specify the column where you want to start (A=1, B=2, C=3, etc.)

  # Insert the data into the Excel sheet, starting from the second row of data
  for row_index, row_data in enumerate(data_to_insert[1:], start=start_row):
      for col_index, value in enumerate(row_data, start=start_col):
          cell = ws.cell(row=row_index, column=col_index, value=value)
          cell.border = border
          cell.alignment = alignment

  # Create a new sheet or use an existing one to insert the DataFrame
  sheet_name2 = 'Unanswered '+current_date_str  # Specify the name of the sheet you want to insert the data into
  ws = workbook[sheet_name2]
  #df1 = pd.DataFrame(df1)

  clean_unanswered = clean_unanswered.astype(str)


  # Convert the DataFrame to a list of lists for inserting into the Excel sheet
  data_to_insert_2 = [clean_unanswered.columns.tolist()] + clean_unanswered.values.tolist()

  # Determine the starting cell where you want to insert the data
  start_row_2 = 7  # Specify the row where you want to start (1-based index)
  start_col_2 = 2  # Specify the column where you want to start (A=1, B=2, C=3, etc.)


  # Insert the data into the Excel sheet
  for row_index, row_data in enumerate(data_to_insert_2[1:], start=start_row_2):
      for col_index, value in enumerate(row_data, start=start_col_2):
          cell = ws.cell(row=row_index, column=col_index, value=value)
          cell.border = border
          cell.alignment = alignment

  ###survey question

  sheet = workbook.create_sheet(title='Survey Questions')
  # Merge cells
  start_cell_6 = "C4"
  end_cell_6 = "D4"
  sheet.merge_cells(start_cell_6 + ":" + end_cell_6)

  # Set content in the merged cell
  sheet[start_cell_6] = "Survey Questions"

  # Apply formatting to the merged cell
  fill = PatternFill(start_color="ffccac", end_color="ffccac", fill_type="solid")
  font = Font(bold=True)  # Bold and gold color
  alignment = Alignment(horizontal="center", vertical="center")
  border = Border(
      left=Side(style="thin"),
      right=Side(style="thin"),
      top=Side(style="thin"),
      bottom=Side(style="thin")
  )



  # Apply the formatting to the merged cell
  for row in sheet.iter_rows(min_row=4, max_row=4, min_col=3, max_col=3):
      for cell in row:
          cell.fill = fill
          cell.font = font
          cell.border = border
          cell.alignment = alignment


  # Set the column widths
  sheet.column_dimensions['D'].width = 123.82


  # Add the header of table
  data_5 = [
      ["1", "How would you rate your understanding of the policy illustration, and product disclosure sheets that you received?"],
      ["2","How would you rate your understanding of the features, and benefits of the product you've purchased, as explained to you by your salesperson?"],
      ["3","How would you rate the professionalism, and politeness of the salesperson?"],
      ["4","How would you rate the salesperson's responsiveness to your request?"]


  ]

  # Set formatting for header of table
  fill = PatternFill(start_color="ffffff",end_color="ffffff",fill_type="solid")
  border = Border(
      left=Side(style="thin"),
      right=Side(style="thin"),
      top=Side(style="thin"),
      bottom=Side(style="thin")
  )


  # Fill the header of table  data and apply formatting
  for row_index, row_data in enumerate(data_5, start=5):
      for col_index, value in enumerate(row_data, start=3):
          cell = sheet.cell(row=row_index, column=col_index)
          cell.value = value
          cell.border = border
          cell.alignment = Alignment(horizontal="left", vertical="center")
          if row_index == 5:
              cell.fill = fill


  # Create a new dataframe for the workbook to be saved in
  workbook_df = pd.DataFrame()
  workbook_df.to_excel(f"/tmp/Hana Call Summary Full Report {current_date_str}.xlsx", index=False)
  workbook.save(f'/tmp/Hana Call Summary Full Report {current_date_str}.xlsx')

  input_file = f'/tmp/Hana Call Summary Full Report {current_date_str}.xlsx'
  output_file = f'/tmp/Hana Call Summary Full Report {current_date_str}.zip'
  password_date = datetime.now().date().strftime('%d%m%Y')
  password = f"HANAETIQA{password_date}"
  
  # Check if the input file exists
  if os.path.exists(input_file):
      try:
          result = subprocess.run(['/opt/bin/7za', 'a', '-p' + password, '-y', output_file, input_file], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
          
          stdout = result.stdout.decode()
          stderr = result.stderr.decode()
          
          print(stdout)
          print(stderr)
  
      except subprocess.CalledProcessError as e:
          print(f"An error occurred: {str(e)}")
  else:
      print(f"Input file '{input_file}' does not exist.")

  new_object_key_excel = f'Hana_Call_Summary_Report/Hana Call Summary Full Report {current_date_str}.xlsx'
  new_object_key_zip = f'Hana_Call_Summary_Report/Hana Call Summary Full Report {current_date_str}.zip'
  
  if os.path.exists(output_file):
    print(f"Output file '{output_file}' successfully created.")
  
  else:
    print(f"Output file '{output_file}' does not exist.")
    
    
  # Upload the locked zip file to S3
  s3_client.upload_file(input_file, bucket_name, new_object_key_excel)
  s3_client.upload_file(output_file, bucket_name, new_object_key_zip)

  if tmr_str is not None:
    schedule_call = clean_unanswered.drop(columns=["Entity"])
    schedule_call.to_excel("/tmp/SMS_Blast_"+tmr_str+".xlsx", index=False)
    input_file = "/tmp/SMS_Blast_"+tmr_str+".xlsx"
    output_file = "/tmp/SMS_Blast_"+tmr_str+".zip"
    
    if os.path.exists(input_file):
      try:
          result = subprocess.run(['/opt/bin/7za', 'a', '-p' + password, '-y', output_file, input_file], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
          
          stdout = result.stdout.decode()
          stderr = result.stderr.decode()
          
          print(stdout)
          print(stderr)
  
      except subprocess.CalledProcessError as e:
          print(f"An error occurred: {str(e)}")
    else:
      print(f"Input file '{input_file}' does not exist.")
  
    if os.path.exists(output_file):
      print(f"SMS BLAST zip file '{output_file}' successfully created.")
    
    else:
      print(f"SMS BLAST zip file '{output_file}' does not exist.")

    SMS_object_key_xlsx = f'Hana_SMS_Blast/SMS_Blast_{tmr_str}.xlsx'
    SMS_object_key = f'Hana_SMS_Blast/SMS_Blast_{tmr_str}.zip'

    s3_client.upload_file(input_file, bucket_name, SMS_object_key_xlsx)
    s3_client.upload_file(output_file, bucket_name, SMS_object_key)